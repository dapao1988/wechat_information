#coding=utf8
"""
# Author: Wenbing.Wang
# Email: wangwenbingood1988@gmail.com
# Created Time : Sun Apr  12 20:59:04 2024

# File Name: example.py
# Description:

"""
import os
import openpyxl
import pprint
from pprint import pprint
import unittest
import re
from wxpy import *
from collections import defaultdict
from collections import OrderedDict
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill, colors
from constants import *

def get_members(group_name):
    print(bot.groups().search(group_name)[0])
    group = bot.groups().search(group_name)[0]
    #使用此方法用来更新群聊成员的详细信息 (地区、性别、签名等)
    group.update_group(members_details=True)
    #获取该群聊组的全部成员对象
    members = group.members
    return members

def organize_member_info(members):
    #member_info_list=defaultdict(list)
    member_info_list=OrderedDict()
    member_info_list[HEAD_INFO] = [ \
            HEAD_NAME, \
            HEAD_NICK_NAME,  \
            HEAD_PROVINCE,  \
            HEAD_CITY, \
            HEAD_SEX, \
            HEAD_SIGNATURE \
    ]

    for index,member in enumerate(members):
        member_name = member.name
        member_nick_name = member.display_name
        member_province = member.province
        member_city = member.city
        #member_sex = "male" if member.sex else "female"
        member_sex = member.sex
        member_signature = member.signature

        member_info_list[index] = [member_name, member_nick_name, member_province, member_city, member_sex, member_signature]


    return member_info_list


def create_or_append_file_with_ui(filename, mode='r'):
    """
    用于查询文件是否存在，以及咨询是否覆盖
    """

    if os.path.exists(filename):
        room_input = "error_state"
        while room_input not in ['yes', 'no']:
            room_input = input(f"The file '{filename}' already exists. Do you want to overwrite it? (yes/no): ").lower()

        if room_input == 'yes':
            mode = 'w'  # 覆盖模式
            ret_filename = filename
        elif room_input == 'no':
            while True:
                ret_filename = input("Please input a new filename: ")
                ret_filename, mode = create_or_append_file_with_ui(ret_filename)
    else:
        mode = 'w'  # 创建新文件
        ret_filename = filename

    return ret_filename, mode

def write_members_to_file(file_name, members_info):
    """
    把相关info写入文件
    """
    ret_filename,mode = create_or_append_file_with_ui(file_name)

    with open(ret_filename, mode) as f:
        for name, info_list in members_info.items():
            f.write(name+':'+'\t'.join(info_list)+'\n')

def write_members_to_excel(file_name, members_info):
    if members_info is None:
        print("the members_info content is empty")
        return

    ret_filename,mode = create_or_append_file_with_ui(file_name)
    filename_without_extention = os.path.splitext(ret_filename)[0]
    assert filename_without_extention, "文件名不能为空"
    total_count = len(members_info)
    writed_count = 0
    attribute_len = len(members_info[HEAD_INFO])

    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(filename_without_extention, index=0);

    for index,(key,value) in enumerate(members_info.items()):
        column_length = len(value)
        assert column_length == attribute_len, f"member unit length({column_length}), not equal with attribute_len({attribute_len})"
        for col in range(1,column_length+1):
            sheet.cell(row=index+1, column=col).value = value[col-1]

        writed_count += 1

    print(f"totally saved {writed_count}/{total_count} items")
    wb.save(ret_filename)

def load_wechat_information(xlsx_filename: str, origin_sheet_name: str) -> dict[int,str]:
    assert xlsx_filename is not None, "xlsx_filename is None!"
    assert origin_sheet_name is not None, "origin_sheet_name is None!"

    wb = openpyxl.load_workbook(xlsx_filename)
    # query the sheet
    #sheet = wb.get_sheet_by_name(origin_sheet_name)
    sheet = wb[origin_sheet_name]
    # set the queried sheet active
    index = wb.index(sheet)
    # activate the sheet for loading data
    wb.active = index

    # loading the data
    row_max = sheet.max_row
    column_max = sheet.max_column
    column_max_letter = get_column_letter(column_max)

    # dict(int, str)
    info_orderdict = OrderedDict()
    count = 0
    # pass the first line for it is head info
    for key,rowOfCellObjects in enumerate(sheet['A2':column_max_letter+f"{row_max}"]):

        #for cellObj in rowOfCellObjects:
        member_name = rowOfCellObjects[0].value
        member_nick_name = rowOfCellObjects[1].value
        member_province = rowOfCellObjects[2].value
        member_city = rowOfCellObjects[3].value
        #member_sex = "male" if member.sex else "female"
        member_sex = rowOfCellObjects[4].value
        member_signature = rowOfCellObjects[5].value

        info_orderdict[key+2] = { \
                HEAD_NAME:member_name, \
                HEAD_NICK_NAME:member_nick_name, \
                HEAD_PROVINCE:member_province, \
                HEAD_CITY:member_city, \
                HEAD_SEX:member_sex, \
                HEAD_SIGNATURE:member_signature \
        }
        count += 1

    print(f"loaded {count}/{row_max-1} items from {xlsx_filename}")

    assert count == row_max-1, f"not all the items in {xlsx_filename} been loaded"

    return info_orderdict


def transform_information(wechat_package_list: dict[int,dict]) -> dict[int,str]:
    # replace some Chinese word, regex for match and sub
    regex_building = re.compile(r'(\w|一|二|三|四|五|六|七|八|九)(号楼|栋)')
    regex_unit = re.compile(r'(\w|一|二)单元')
    # regex last! last! last!, for special description
    regex_get_building_unit_room = re.compile('((\d)(- |-|—| -| —|一| －|–|－| - |--|  )){1}((\d)(- |-|—| -| —|一| －|–|－| - |--|   ))?(\d+)')

    def chinese_to_arabic(character):
       return chinese_to_arabic_dict[character] if not character.isdigit() else character

    def refract_regex_string(building_string, unit_string, room_string) -> str:
        assert building_string is not None, "building_string is None"
        assert room_string is not None, "room_string is None"

        # 如果没备注几单元，默认为1单元
        if unit_string is None:
            unit_string = '1'

        return f"{building_string}-{unit_string}-{room_string}"

    def is_single_digit(character):
            return len(character) == 1 and character.isdigit()

    transform_dict = defaultdict(str)

    for key, member_info in wechat_package_list.items():
        value = member_info["name"]
        print(f"{key}:{value}")

        assert value is not None, "room info is null"

        # 检查楼栋替换
        match = regex_building.search(value)
        if match:
            character = match.group(1)
            character = chinese_to_arabic(character)
            building_replaced_string = regex_building.sub(character+'-', value)
        else:
            building_replaced_string = value

        # 检查单元替换, 检查不到自动不做sub替换
        match = regex_unit.search(building_replaced_string)
        if match:
            character = match.group(1)
            character = chinese_to_arabic(character)
            unit_replaced_string = regex_unit.sub(character+'-', building_replaced_string)
        else:
            unit_replaced_string = building_replaced_string

        # 捞取'楼栋-单元-户号'
        match = regex_get_building_unit_room.search(unit_replaced_string)
        if match:
            building = match.group(2)
            unit = match.group(5)
            # 如果没写单元号，则默认置1
            if unit is None:
                unit = 1
            room = match.group(7)
            # 如果是单个数字，则改为901，比如1-1-9
            if is_single_digit(room):
                room = str(room)+'01'
            building_unit_room = "-".join(map(str, [building, unit, room]))
        else:
            assert False, "It must be match 楼栋-单元-户"

        #building_unit_room = regex_get_building_unit_room.sub(r'\2-\5-\7', unit_replaced_string)

        transform_dict[key] = building_unit_room

        print(f'{value} -> {building_unit_room}')

    return transform_dict

def classify_residents(wechat_package_transform_dict: dict[int,str]):

    def custom_sort(item):
        key, value = item
        building, unit, room = map(int, value.split("-"))
        return (building, unit, room, key)

    # 按照自定义排序函数对字典的键值对进行排序
    sorted_user_dict = sorted(wechat_package_transform_dict.items(), key=custom_sort)

    pprint(sorted_user_dict)

    return sorted_user_dict

def copy_wechat_orginal_sheet(source_file, target_file, source_sheet, ):
    if source_file is None or \
            target_file is None or \
            source_sheet is None:
        assert False, "input parameter is None"
    # 打开源文件 A.xlsx 和目标文件 B.xlsx
    source_wb = openpyxl.load_workbook(filename=source_file)
    #target_wb = openpyxl.load_workbook(filename=target_file)
    target_wb = openpyxl.Workbook()

    # 获取源文件和目标文件的活动工作表
    source_sheet = source_wb[source_sheet]
    # 获取默认标签的第一个sheet
    target_sheet = target_wb['Sheet']

    # 复制源文件的活动工作表到目标文件
    target_sheet.title = source_sheet.title  # 修改目标文件的工作表名称为源文件的名称
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

    # 保存目标文件
    #target_wb.save(target_file)
    return target_wb

def get_sex(sex):
    """
    # 男性
    MALE = 1
    # 女性
    FEMALE = 2
    """
    sex_str = ""
    sex_integer = int(sex)

    if sex_integer == 0:
        sex_str = "未知"
    elif sex_integer == 1:
        sex_str = "男"
    elif sex_integer == 2:
        sex_str = "女"
    else:
        assert False, "sex value is error!!!"

    return sex_str

def save_wechat_information(source_file, file_name,  source_sheet, building_unit_room_dict, info_orderdict) -> None:
    wb = copy_wechat_orginal_sheet(source_file, file_name, source_sheet)
    ret_filename,mode = create_or_append_file_with_ui(file_name)
    total_nubmer = len(building_unit_room_dict)
    print(f"total user: {total_nubmer}")
    # 创建一个新的 Excel 工作簿
    #wb = openpyxl.Workbook()

    # 遍历排序后的字典，按照楼栋写入不同的 sheet
    current_building = 0
    current_sheet = None
    total_building = 0
    total_computed_number = 0
    # 每一栋的住户数
    count_building_unit = defaultdict(int)

    # 统计一个家庭的人数
    family_count = 0
    # 上一行的用户"楼栋-单元-户"信息字符串
    value_pre = ""

    for key, value in building_unit_room_dict:

        building, _, _ = value.split("-")
        if building != current_building:
            current_building = building
            current_sheet = wb.create_sheet(title=f"{building}栋")
            current_sheet.append(["户号", "首页表行号", "群昵称", "省份", "城市", "性别", "签名"])

            total_building += 1
            # reset family count
            family_count = 0

        # 计算一栋楼的人数
        count_building_unit[building] += 1
        # 将正则过后的"楼栋-单元-户"信息追加到excel sheet中, 同时把首页中的用户其他信息也捞取过来一并写入
        # key, 为excel第一个sheet中的行号，是用户信息唯一标志
        current_sheet.append([value, key, \
                info_orderdict[key][HEAD_NICK_NAME], \
                info_orderdict[key][HEAD_PROVINCE], info_orderdict[key][HEAD_CITY], \
                get_sex(info_orderdict[key][HEAD_SEX]), \
                info_orderdict[key][HEAD_SIGNATURE]])
        # 统计总分析户数，便于和首页中捞取的原信息做对比，避免有遗漏
        total_computed_number += 1

        # 先确定当前字符串是否是同一户，便于标记颜色和字体
        if value == value_pre:
            family_count += 1
        else:
            # 信息改变时，先回溯设置超过3个人的家庭单位
            if family_count >= WARNING_FAMILY_NUBMER:
                # 设置字体和填充颜色
                #fontObj1 = Font(name='Times New Roman', bold=True)
                #bold_itatic_24_font = Font(name='微软雅黑', size=24, italic=False, color=colors.RED, bold=False)
                #sht["A1"].fill = PatternFill("solid", fgColor="1874CD")
                #styleObj1 = Style(font=fontObj1)

                font = Font(name='微软雅黑',  bold=True)  # 设置红色字体
                fill = PatternFill(fill_type="solid", start_color="FFEC8B")  # 设置黄色填充
                current_row_index = current_sheet.max_row

                # 循环向前设置多行的字体和颜色属性
                for i in range(current_row_index-1, current_row_index-family_count-1, -1):
                    for cell in current_sheet[i]:
                        cell.font = font
                        cell.fill = fill

            # 最后将计数变量归位, 当前用户也算一个，所以为1
            family_count = 1

        # 保存上一条"楼栋-单元-户"信息字符串
        value_pre = value

    print(f"total_computed_number:{total_computed_number}, total_building:{total_building}")

    # 最后新建一个sheet用于存储楼栋人数分布
    sheet = wb.create_sheet(title="楼栋人数分布")
    sheet.append(["楼栋号", "人数"])
    for build_num,value in count_building_unit.items():
        print(f"{build_num}:{value}")
        sheet.append([build_num, value])

    # 保存 Excel 文件
    wb.save(ret_filename)
    return

class TestWechatMembers(unittest.TestCase):
    def test_transform_information(self):
        room1 = "1-1-1602 桔子撞墙"
        room2 = "3号楼2单元402"
        room3 = "1—1—9"
        room4 = "1—1—902"
        room5 = "1号楼二单元1501"
        room6 = "2-2-701贴地飞行"
        room7 = "1 —1402"
        room8 = "2栋一单元801"
        room9 = "1—2—2402"
        room10 = "白合2-1-1101"
        room11 = "1一2-2301"
        room12 = "Lillian🐰5-1-902"
        room13 = "9 －1 －1302"
        room14 = "4-1-1201wewe"
        room15 = "8—1902"
        room16 = "8-1–1002 老范"
        room17 = "7—1—902"
        room18 = "7-1302bw."
        room19 = "9—1702"
        room20 = "2－2－1101"

        # 创建一个空字典, 存放最开始的用户数据信息
        room_dict = {}

        # 遍历每个用户字符串
        for i in range(1, 21):  # 假设有19个用户字符串
            room = locals()["room" + str(i)]  # 根据变量名获取变量的值
            room_dict[i] = room  # 将字符串作为值添加到字典中

        # 打印结果
        pprint(room_dict)
        source_filename = "光谷龙山湖业主2群.xlsx"
        source_filename_sheet = "光谷龙山湖业主2群"
        target_filename =  "test.xlsx"
        info_orderdict = load_wechat_information(source_filename, source_filename_sheet)
        #transform_dict = transform_information(room_dict)
        transform_dict = transform_information(info_orderdict)
        classified_dict = classify_residents(transform_dict)
        #copy_wechat_orginal_sheet(source_filename, target_filename, source_filename_sheet)
        save_wechat_information(source_filename, target_filename, source_filename_sheet, classified_dict, info_orderdict)


if __name__ == '__main__':
    unittest.main()
