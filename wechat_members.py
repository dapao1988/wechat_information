#coding=utf8
"""
# Author: Wenbing.Wang
# Email: wangwenbingood1988@gmail.com
# Created Time : Sun Apr  7 11:32:04 2024

# File Name: example.py
# Description:

"""
import os
import openpyxl
from wxpy import *
#from collections import defaultdict
from collections import OrderedDict
from constants import *

def get_members(group_name):
    print(bot.groups().search(group_name)[0])
    group = bot.groups().search(group_name)[0]
    #使用此方法用来更新群聊成员的详细信息 (地区、性别、签名等)
    group.update_group(members_details=True)
    #获取该群聊组的全部成员对象
    members = group.members
    return members

def organize_member_info(members):
    #member_info_list=defaultdict(list)
    member_info_list=OrderedDict()
    member_info_list[HEAD_INFO] = [ \
            HEAD_NAME, \
            HEAD_NICK_NAME, \
            HEAD_PROVINCE, \
            HEAD_CITY, \
            HEAD_SEX, \
            HEAD_SIGNATURE \
    ]

    for index,member in enumerate(members):
        member_name = member.name
        member_nick_name = member.display_name
        member_province = member.province
        member_city = member.city
        #member_sex = "male" if member.sex else "female"
        member_sex = member.sex
        member_signature = member.signature

        member_info_list[index] = [member_name, member_nick_name, member_province, member_city, member_sex, member_signature]


    return member_info_list

def write_members_to_excel(file_name, members_info):
    if members_info is None:
        print("the members_info content is empty")
        return

    ret_filename,mode = create_or_append_file_with_ui(file_name)
    filename_without_extention = os.path.splitext(ret_filename)[0]
    assert filename_without_extention, "文件名不能为空"
    total_count = len(members_info)
    writed_count = 0
    attribute_len = len(members_info[HEAD_INFO])

    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(filename_without_extention, index=0);

    for index,(key,value) in enumerate(members_info.items()):
        column_length = len(value)
        assert column_length == attribute_len, f"member unit length({column_length}), not equal with attribute_len({attribute_len})"
        for col in range(1,column_length+1):
            sheet.cell(row=index+1, column=col).value = value[col-1]

        writed_count += 1

    print(f"totally saved {writed_count}/{total_count} items")
    wb.save(ret_filename)

def create_or_append_file_with_ui(filename, mode='r'):

    if os.path.exists(filename):
        user_input = "error_state"
        while user_input not in ['yes', 'no']:
            user_input = input(f"The file '{filename}' already exists. Do you want to overwrite it? (yes/no): ").lower()

        if user_input == 'yes':
            mode = 'w'  # 覆盖模式
            ret_filename = filename
        elif user_input == 'no':
            while True:
                ret_filename = input("Please input a new filename: ")
                ret_filename, mode = create_or_append_file_with_ui(ret_filename)
    else:
        mode = 'w'  # 创建新文件
        ret_filename = filename

    return ret_filename, mode

def write_members_to_file(file_name, members_info):
    ret_filename,mode = create_or_append_file_with_ui(file_name)

    with open(ret_filename, mode) as f:
        for name, info_list in members_info.items():
            f.write(name+':'+'\t'.join(info_list)+'\n')



if __name__ == '__main__':
    # 初始化机器人，扫码登陆
    bot = Bot()

    #成想要统计的群聊名字，很久没有聊天记录的群最好先发条消息
    group_n = "光谷龙山湖业主2群"
    members = get_members(group_n)
    members_info = organize_member_info(members)
    write_members_to_excel(group_n+'.xlsx', members_info)
